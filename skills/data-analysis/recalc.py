#!/usr/bin/env python3
"""
Excel 公式重算脚本

核心思路：
1. 通过写入 LibreOffice Basic 宏，让 `soffice` 在无界面模式下执行“全量重算+保存”。
2. 重算后用 openpyxl 读取结果（`data_only=True`），扫描单元格中的 Excel 错误值。
3. 再读取一次工作簿（`data_only=False`）统计公式总数，作为结果上下文。
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def setup_libreoffice_macro():
    """
    准备 LibreOffice 宏（若不存在则创建）。

    该函数会确保 `Standard/Module1.xba` 中存在 `RecalculateAndSave` 过程，
    供后续通过 `vnd.sun.star.script:` URL 直接调用。
    """
    # 不同系统下 LibreOffice 用户配置目录不同
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    # 若宏文件已经存在且包含目标过程名，直接复用
    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True

    # 首次运行时目录可能不存在：先触发一次 LibreOffice 初始化，再创建目录
    if not os.path.exists(macro_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'],
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)

    # 写入一个最小可用宏：
    # - calculateAll(): 全工作簿重算
    # - store(): 保存当前文档
    # - close(True): 关闭文档并确认保存
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        # 直接覆盖写入，确保脚本内容与预期一致
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        # 此处只回传布尔值，让上层统一返回 JSON 错误结构
        return False


def recalc(filename, timeout=30):
    """
    重算 Excel 中的公式，并报告错误单元格。

    Args:
        filename: Excel 文件路径
        timeout: 重算最大等待时间（秒）

    Returns:
        dict:
            - 成功时包含 status / total_errors / total_formulas / error_summary
            - 失败时包含 error
    """
    # 先做输入文件存在性校验，避免后续 soffice 报错信息不直观
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    # 统一使用绝对路径，避免 soffice 在不同 cwd 下找不到文件
    abs_path = str(Path(filename).absolute())

    # 依赖宏执行重算，宏不可用则直接返回
    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}

    # 使用 LibreOffice 脚本 URL 调用应用级宏（location=application）
    # 参数顺序：先脚本 URL，再文档路径
    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    # 处理 Linux / macOS 的 timeout 命令差异：
    # - Linux 通常是 timeout
    # - macOS 原生命令没有 timeout，常见是 coreutils 提供的 gtimeout
    # Windows 分支不加 timeout 包裹（保留原行为）
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # 检测 macOS 是否安装 gtimeout
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        # 若找到了 timeout 包装器，则整体命令前置 timeout
        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd

    # 执行重算命令；text=True 方便按字符串处理 stderr
    result = subprocess.run(cmd, capture_output=True, text=True)

    # 约定：124 为 timeout 命令超时退出码（不在此处直接判为硬错误）
    # 其他非零退出码视为异常，尝试给出可读错误信息
    if result.returncode != 0 and result.returncode != 124:  # 124 is timeout exit code
        error_msg = result.stderr or 'Unknown error during recalculation'
        # 若信息中提及 Module1 或看不到过程名，通常说明宏没有按预期注册/加载
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly'}
        else:
            return {'error': error_msg}

    # 重算后读取结果值，扫描 Excel 错误类型（全表遍历，不限范围）
    try:
        wb = load_workbook(filename, data_only=True)

        # 常见 Excel 错误码集合
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        # 为每种错误记录出现位置（sheet!cell）
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # iter_rows() 会覆盖当前工作表已使用区域内的所有单元格
            for row in ws.iter_rows():
                for cell in row:
                    # 仅对字符串做错误匹配；数字、日期等类型跳过
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            # 使用包含判断而不是全等：兼容单元格中带补充文字的情况
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                # 命中一个错误码后立即停止当前单元格的后续匹配
                                break

        wb.close()

        # 构建返回结果摘要
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }

        # 仅输出有命中的错误类型；每类最多展示 20 个位置，避免 JSON 过大
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }

        # 额外统计公式总数（再次读取 data_only=False 才能拿到原始公式字符串）
        # 这里单独再开一个 workbook，避免与上面的 data_only 语义混用
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    # openpyxl 中公式通常以 '=' 开头的字符串形式存在
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()

        result['total_formulas'] = formula_count

        return result

    except Exception as e:
        # 捕获读取/解析期异常并统一输出
        return {'error': str(e)}


def main():
    # CLI 参数不足时给出用法说明并返回非零退出码
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    # 可选超时参数，默认 30 秒
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    # 统一以 JSON 输出，便于脚本间调用
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
